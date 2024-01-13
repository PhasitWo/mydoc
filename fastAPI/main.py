from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import MergedCell
from fastapi import FastAPI
from pydantic import BaseModel
import uvicorn

app = FastAPI()


@app.get("/")
async def root():
    return {"message": "Hello World"}


class Req(BaseModel):
    formPath: str
    saveAt: str
    input: dict
    keyword: dict


@app.post("/write_form/")
async def write_form(request: Req):
    try:
        writeForm(request.formPath, request.saveAt, request.input, request.keyword)
    except:
        return {"status": "FAIL"}
    return {"status": "SUCCESS"}


def writeForm(formPath, saveAt, input, keyword):
    workbook = load_workbook(formPath)
    sheet_name = workbook.sheetnames
    for name in sheet_name:
        sheet = workbook[name]
        for i in range(1, sheet.max_column + 1):
            for j in range(1, sheet.max_row + 1):
                cell = sheet[get_column_letter(i) + str(j)]
                if isinstance(cell, MergedCell):
                    continue
                if cell.value == None or "=" in cell.value:
                    continue
                value = str(cell.value)
                for prop in keyword:
                    replacement = (
                        keyword[prop]["default"] if input[prop] == "" else input[prop]
                    )
                    value = value.replace(keyword[prop]["key"], str(replacement))
                if value.isnumeric():
                    value = float(value)
                sheet[get_column_letter(i) + str(j)] = value
    workbook.save(saveAt)
    workbook.close()


if __name__ == "__main__":
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
    )
