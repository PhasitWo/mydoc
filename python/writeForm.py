import json
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import MergedCell
import os


def writeForm(formPath, saveAt, input, keyword):
    workbook = load_workbook(formPath)
    sheet_name = workbook.sheetnames
    for name in sheet_name:
        sheet = workbook[name]
        for i in range(1, sheet.max_column + 1):
            for j in range(1, sheet.max_row + 1):
                cell = sheet[get_column_letter(i) + str(j)]
                if isinstance(cell, MergedCell):
                    continue
                if cell.value == None or "=" in cell.value:
                    continue
                value = str(cell.value)
                for prop in keyword:
                    replacement = (
                        keyword[prop]["default"] if input[prop] == "" else input[prop]
                    )
                    value = value.replace(keyword[prop]["key"], str(replacement))
                if value.isnumeric():
                    value = float(value)
                sheet[get_column_letter(i) + str(j)] = value
    workbook.save(saveAt)
    workbook.close()


if __name__ == "__main__":
    try:
        with open("middleware.json", "r", encoding="UTF-8") as openfile:
            inp = json.load(openfile)
            writeForm(inp["formPath"], inp["saveAt"], inp["input"], inp["keyword"])
            print("SUCCESS", end="")
    except:
        print("FAIL", end="")
    finally:
        os.remove("middleware.json")
